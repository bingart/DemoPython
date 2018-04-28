#!/usr/bin/python
# coding=utf-8

import time
import os
import re
from http_helper import HttpHelper
from url_helper import UrlHelper
from str_helper import StrHelper
from nlp_helper import NLPHelper
from crypt_helper import CryptHelper
from mongo_helper import MongoHelper
from file_helper import FileHelper

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

MONGO_HOST = "172.16.40.140"
MONGO_DATABASE_NAME = "ZDBBingCom"

BING_HOST = "https://www.bingapis.com"
BING_APP_ID = "F7E7AFFA19051F569B9724DEE37BE08A70F468EB"
# https://www.bingapis.com/api/v5/search?appid=F7E7AFFA19051F569B9724DEE37BE08A70F468EB&q=health&offset=0&count=20&mkt=en-US

def loadKeywords():
    try: 
        wb = load_workbook('D:/work/Traffic/doc/Keyword/BingKeyword.xlsx')
        sheetNames = wb.get_sheet_names()
        print (sheetNames)
        ws = wb.get_sheet_by_name("Sheet1")
        colSet = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'K']
        keyList = []
        dict = {}
        for row in range(1, 27):
            for col in colSet:
                cname = col + str(row)
                cell = ws[cname]
                value = cell.value
                if value != None and len(value) > 0:
                    if not value.lower() in dict:
                        dict[value.lower()] = value.lower()
                        key = {
                            'index': row,
                            'title': value,
                            'level': 0,
                            'state': 'CREATED'
                            }
                        keyList.append(key)

        collection.insertMany(keyList)
    except Exception as err :
        print(err)
        
if __name__=="__main__":
    print("main")
    loadKeywords()
    print("exit")
    
    
    
    
    